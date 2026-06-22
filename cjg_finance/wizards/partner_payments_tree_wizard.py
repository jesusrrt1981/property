# -*- coding: utf-8 -*-
"""
Wizard para Listado de Abonos del Cliente sin Inicial (vista tree + XLSX).

Equivalente a ``listado_abonos_cliente.php`` de Testarossa
(caja/view/cliente/, 58 líneas). Mismo concepto que el listado de
abonos iniciales pero excluyendo el movimiento ``ini``.

Columnas del legacy:
    Fecha, Docto, TC, Serie, No. Doc, Monto, Caja.

A diferencia del wizard PDF existente, este entrega:
    * Botón "Ver Listado (tree)" → vista tree filtrada de cjg.pos.payment.receipt
    * Botón "Exportar XLSX"     → genera un XLSX con openpyxl
"""
from io import BytesIO

from odoo import api, fields, models, _


class PartnerPaymentsTreeWizard(models.TransientModel):
    _name = "cjg.finance.partner.payments.tree.wizard"
    _description = "Listado de Abonos del Cliente (sin Inicial, tree + XLSX)"

    partner_id = fields.Many2one(
        "res.partner", string="Cliente", required=True,
    )
    date_from = fields.Date(string="Fecha Desde")
    date_to = fields.Date(string="Fecha Hasta")
    company_id = fields.Many2one(
        "res.company", string="Compañía",
        default=lambda self: self.env.company,
    )
    state = fields.Selection(
        selection=lambda self: self._get_state_selection(),
        string="Estado",
        default="active",
    )

    @api.model
    def _get_state_selection(self):
        return [
            ("active", "Solo Activos"),
            ("cancelled", "Solo Anulados"),
            ("all", "Todos (incluye anulados)"),
        ]

    def _build_domain(self):
        """Construye el dominio sobre ``cjg.pos.payment.receipt``.

        Equivalente conceptual al $WHERE del legacy
        ``listado_abonos_cliente.php`` (excluye ini).
        """
        self.ensure_one()
        domain = [
            ("partner_id", "=", self.partner_id.id),
            ("movement_type", "!=", "ini"),
        ]
        if self.state == "active":
            domain.append(("state", "!=", "cancelled"))
        elif self.state == "cancelled":
            domain.append(("state", "=", "cancelled"))
        if self.date_from:
            domain.append(("date", ">=", self.date_from))
        if self.date_to:
            domain.append(("date", "<=", self.date_to))
        if self.company_id:
            domain.append(("company_id", "=", self.company_id.id))
        return domain

    def action_open_tree(self):
        """Abre la vista tree de ``cjg.pos.payment.receipt`` con el filtro aplicado."""
        self.ensure_one()
        domain = self._build_domain()
        return {
            "name": _("Abonos del Cliente — %s") % (self.partner_id.name or ""),
            "type": "ir.actions.act_window",
            "res_model": "cjg.pos.payment.receipt",
            "view_mode": "tree,form",
            "domain": domain,
            "context": {
                "create": False,
            },
        }

    def action_export_xlsx(self):
        """Genera un XLSX con los abonos sin inicial filtrados."""
        self.ensure_one()
        domain = self._build_domain()
        receipts = self.env["cjg.pos.payment.receipt"].search(
            domain, order="date asc, name asc",
        )
        if not receipts:
            return {
                "type": "ir.actions.client", "tag": "display_notification",
                "params": {
                    "title": _("Sin abonos"),
                    "message": _("No hay abonos (sin inicial) con esos filtros."),
                    "type": "warning",
                },
            }

        xlsx_bytes = self._build_xlsx(receipts)
        attachment = self.env["ir.attachment"].create({
            "name": "Abonos_Cliente_%s.xlsx" % (
                self.partner_id.name or "cliente",
            ),
            "datas": xlsx_bytes,
            "res_model": self._name,
            "res_id": self.id,
            "type": "binary",
            "mimetype": "application/vnd.openxmlformats-officedocument."
                        "spreadsheetml.sheet",
        })

        return {
            "type": "ir.actions.act_url",
            "url": "/web/content/%s?download=true" % attachment.id,
            "target": "new",
        }

    def _build_xlsx(self, receipts):
        """Construye el XLSX en memoria usando openpyxl."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Abonos Cliente"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="2F5496", end_color="2F5496", fill_type="solid",
        )
        center = Alignment(horizontal="center", vertical="center")

        cols = [
            ("Fecha", 12),
            ("N° Docto", 18),
            ("Tipo Mov.", 16),
            ("Serie", 8),
            ("No. Doc", 16),
            ("Tipo Cambio", 14),
            ("Caja", 18),
            ("Monto RD$", 16),
            ("Estado", 14),
        ]
        for col_idx, (label, width) in enumerate(cols, start=1):
            cell = ws.cell(row=1, column=col_idx, value=label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        for row_idx, r in enumerate(receipts, start=2):
            ws.cell(row=row_idx, column=1,
                    value=r.date.strftime("%Y-%m-%d") if r.date else "")
            ws.cell(row=row_idx, column=2, value=r.name or "")
            ws.cell(row=row_idx, column=3,
                    value=r.concept_id.name if r.concept_id else (r.movement_type or ""))
            ws.cell(row=row_idx, column=4, value=r.serie or "")
            ws.cell(row=row_idx, column=5, value=r.ref or "")
            ws.cell(row=row_idx, column=6, value=r.foreign_currency_id.name or "")
            ws.cell(row=row_idx, column=7,
                    value=r.cashbox_id.name if r.cashbox_id else "")
            ws.cell(row=row_idx, column=8, value=r.amount_paid or 0.0)
            ws.cell(row=row_idx, column=9, value=dict(r._fields["state"].selection)
                    .get(r.state, r.state or ""))

        total = sum(r.amount_paid or 0.0 for r in receipts)
        total_row = len(receipts) + 2
        ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=total_row, column=8, value=total).font = Font(bold=True)

        ws.freeze_panes = "A2"

        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()
