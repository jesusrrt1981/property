# -*- coding: utf-8 -*-
"""
Wizard para Listado de Abonos Iniciales del Cliente (vista tree + export XLSX).

Equivalente a ``listado_abono_inicial.php`` (vista HTML <table>) de Testarossa
(caja/view/cliente/, 99 líneas). El legacy filtra movimientos tipo
'ini' (Pago Inicial) para un partner dado y muestra las columnas:

    Fecha, Docto, TC (tipo de cambio), Tipo Doc, Serie, No. Doc, Monto

Adicional al wizard PDF existente (``partner_initial_payments_wizard.py``)
que sólo imprime un reporte, este wizard abre una vista ``tree`` filtrable
sobre los mismos recibos y permite exportar el resultado a XLSX.

Mismas reglas de filtrado que el wizard PDF, pero la salida es:
    * Botón "Ver Listado (tree)" → ir a la vista tree filtrada de cjg.pos.payment.receipt
    * Botón "Exportar XLSX"     → genera un XLSX con openpyxl
"""
from io import BytesIO

from odoo import api, fields, models, _
from odoo.exceptions import UserError


class PartnerInitialPaymentsTreeWizard(models.TransientModel):
    _name = "cjg.finance.partner.initial.payments.tree.wizard"
    _description = "Listado de Abonos Iniciales del Cliente (tree + XLSX)"

    partner_id = fields.Many2one(
        "res.partner", string="Cliente", required=True,
    )
    date_from = fields.Date(string="Fecha Desde")
    date_to = fields.Date(string="Fecha Hasta")
    company_id = fields.Many2one(
        "res.company", string="Compañía",
        default=lambda self: self.env.company,
    )
    state = fields.Selection(
        selection=lambda self: self._get_state_selection(),
        string="Estado",
        help="Filtra por estado del recibo. Por defecto sólo activos.",
        default="active",
    )

    @api.model
    def _get_state_selection(self):
        return [
            ("active", "Solo Activos"),
            ("cancelled", "Solo Anulados"),
            ("all", "Todos (incluye anulados)"),
        ]

    def _build_domain(self):
        """Construye el dominio sobre ``cjg.pos.payment.receipt``.

        Equivalente conceptual al $WHERE del legacy
        ``listado_abono_inicial.php``.
        """
        self.ensure_one()
        domain = [
            ("partner_id", "=", self.partner_id.id),
            ("movement_type", "=", "ini"),
        ]
        if self.state == "active":
            domain.append(("state", "!=", "cancelled"))
        elif self.state == "cancelled":
            domain.append(("state", "=", "cancelled"))
        # state == 'all' no agrega filtro
        if self.date_from:
            domain.append(("date", ">=", self.date_from))
        if self.date_to:
            domain.append(("date", "<=", self.date_to))
        if self.company_id:
            domain.append(("company_id", "=", self.company_id.id))
        return domain

    def action_open_tree(self):
        """Abre la vista tree de ``cjg.pos.payment.receipt`` con el filtro aplicado.

        Equivalente a hacer click en el botón "Listar" del legacy.
        """
        self.ensure_one()
        domain = self._build_domain()
        return {
            "name": _("Abonos Iniciales — %s") % (self.partner_id.name or ""),
            "type": "ir.actions.act_window",
            "res_model": "cjg.pos.payment.receipt",
            "view_mode": "tree,form",
            "domain": domain,
            "context": {
                "search_default_group_by_movement_type": 0,
                "create": False,
            },
        }

    def action_export_xlsx(self):
        """Genera un XLSX con los abonos iniciales filtrados."""
        self.ensure_one()

        domain = self._build_domain()
        receipts = self.env["cjg.pos.payment.receipt"].search(
            domain, order="date asc, name asc",
        )
        if not receipts:
            return {
                "type": "ir.actions.client", "tag": "display_notification",
                "params": {
                    "title": _("Sin abonos"),
                    "message": _("No hay abonos iniciales con esos filtros."),
                    "type": "warning",
                },
            }

        xlsx_bytes = self._build_xlsx(receipts)
        attachment = self.env["ir.attachment"].create({
            "name": "Abonos_Iniciales_%s.xlsx" % (
                self.partner_id.name or "cliente",
            ),
            "datas": xlsx_bytes,
            "res_model": self._name,
            "res_id": self.id,
            "type": "binary",
            "mimetype": "application/vnd.openxmlformats-officedocument."
                        "spreadsheetml.sheet",
        })

        return {
            "type": "ir.actions.act_url",
            "url": "/web/content/%s?download=true" % attachment.id,
            "target": "new",
        }

    def _build_xlsx(self, receipts):
        """Construye el XLSX en memoria usando openpyxl.

        Columnas basadas en el legacy listado_abono_inicial.php SELECT.
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Abonos Iniciales"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="2F5496", end_color="2F5496", fill_type="solid",
        )
        center = Alignment(horizontal="center", vertical="center")

        cols = [
            ("Fecha", 12),
            ("N° Docto", 18),
            ("Tipo Doc", 16),
            ("Serie", 8),
            ("No. Doc", 16),
            ("Tipo Cambio", 14),
            ("Moneda Orig.", 14),
            ("Monto Orig.", 16),
            ("Monto RD$", 16),
            ("Estado", 14),
        ]
        for col_idx, (label, width) in enumerate(cols, start=1):
            cell = ws.cell(row=1, column=col_idx, value=label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        for row_idx, r in enumerate(receipts, start=2):
            ws.cell(row=row_idx, column=1, value=r.date.strftime("%Y-%m-%d") if r.date else "")
            ws.cell(row=row_idx, column=2, value=r.name or "")
            ws.cell(row=row_idx, column=3, value=dict(r._fields["document_type"].selection)
                    .get(r.document_type, r.document_type or ""))
            ws.cell(row=row_idx, column=4, value=r.serie or "")
            ws.cell(row=row_idx, column=5, value=r.ref or "")
            ws.cell(row=row_idx, column=6, value=r.foreign_currency_id.name or "")
            ws.cell(row=row_idx, column=7, value=r.foreign_currency_id.symbol or "")
            ws.cell(row=row_idx, column=8, value=r.amount_currency or 0.0)
            ws.cell(row=row_idx, column=9, value=r.amount_paid or 0.0)
            ws.cell(row=row_idx, column=10, value=dict(r._fields["state"].selection)
                    .get(r.state, r.state or ""))

        total = sum(r.amount_paid or 0.0 for r in receipts)
        total_row = len(receipts) + 2
        ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=total_row, column=9, value=total).font = Font(bold=True)

        ws.freeze_panes = "A2"

        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()
